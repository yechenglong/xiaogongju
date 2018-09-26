import openpyxl

def xiugai(j,ruanjian,w):


    if "步进" in w or "信号" in w or "¶" in w:
        w = w.replace("步进", "")
        w = w.replace("信号", "")
        w = w.replace("¶", "")
    if "//" in w:
        pos = w.index("//")
        w = w[:pos]

    # 写连接的测试点
    H = w + "连接上"+ruanjian
    ws2["H" + str(j)] = H
    J = "1.将" + w + "连接\n" + "2.查看"+ruanjian
    ws2["J" + str(j)] = J
    K = ruanjian+"上有相应的显示或者日志记录"
    ws2["K" + str(j)] = K

    # 写断开连接的测试点
    H = w + "和"+ruanjian+"断开连接"
    ws2["H" + str(j + 1)] = H
    J = "1.将" + w + "的IO断开连接"
    ws2["J" + str(j + 1)] = J
    K = "报错，说明是" + w + "断开连接"
    ws2["K" + str(j + 1)] = K

    # 写断开连接重连的测试点
    H = w + "和"+ruanjian+"断开连接后重新连接"
    ws2["H" + str(j + 2)] = H
    J = "1.将" + w + "的IO断开连接\n" + "2.将" + w + "的IO连接"
    ws2["J" + str(j + 2)] = J
    K = "可重新连接"
    ws2["K" + str(j + 2)] = K


wb1 = openpyxl.load_workbook("F:/work/IA988/060集成测试用例/IO表.xlsx")
wb = openpyxl.Workbook()
wb2 = openpyxl.load_workbook("F:/work/IA988/060集成测试用例/test.xlsx")
ws2 = wb2["Sheet3"]
j = 2
for sheet in wb1.sheetnames:
    ws1 = wb1[sheet]
    if sheet == 'AA1' or sheet == 'AA2':
        for row in range(3, ws1.max_row + 1):
            w = ws1.cell(row=row, column=5).value
            if w != None:
                if row == 30 or row ==31:
                    continue
                xiugai(j,sheet,w)
                j += 3

    else:
        for row in range(3, ws1.max_row + 1):
            w = ws1.cell(row=row, column=5).value
            if w != None:
                xiugai(j, 'PLC', w)
                j += 3

wb2.save("F:/work/IA988/060集成测试用例/test.xlsx")











