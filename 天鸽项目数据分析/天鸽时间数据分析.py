import openpyxl
import datetime

import uiautomation as automation



def string_toDatetime(st):
    return datetime.datetime.strptime(st,"%Y-%m-%d %H:%M:%S.%f")

#python3输入
path = input('请输入文件目录\n')
#python27输入
# print(u'请输入文件目录')
# path = raw_input()
wb = openpyxl.load_workbook(path)
ws = wb["Sheet1"]

for row in range(2, ws.max_row + 1):
    end1 = ws.cell(row=row, column=6).value
    start1 =ws.cell(row=row, column=5).value
    if start1 == None:
        finaltime = "no start time"
        ws.cell(row=row, column=7).value = finaltime
    else:
        if end1 == None:
            end1_time = string_toDatetime(ws.cell(row=row+1, column=5).value)
            start1_time = string_toDatetime(ws.cell(row=row, column=5).value)
            time1 = (end1_time - start1_time)
            finaltime = str(time1.seconds) + "." + str(time1.microseconds)[:3]
        else:
            end1_time = string_toDatetime(ws.cell(row=row, column=6).value)
            start1_time = string_toDatetime(ws.cell(row=row, column=5).value)
            time1 = (end1_time - start1_time)
            finaltime = str(time1.seconds)+"."+str(time1.microseconds)[:3]
    print(finaltime)
    ws.cell(row=row, column=7).value = finaltime
wb.save(path)





